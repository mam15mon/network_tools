from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
                           QTextEdit, QPushButton, QButtonGroup, QRadioButton,
                           QTableWidget, QTableWidgetItem, QHeaderView,
                           QFileDialog, QMessageBox)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os

from src.utils.nat_parser import NATParser
class NatParserTab(QWidget):
    def __init__(self):
        super().__init__()
        self.current_data = []
        self.failed_entries = []
        self.current_device_type = None
        self.setup_ui()

    def setup_ui(self):
        """设置UI界面"""
        layout = QVBoxLayout(self)

        # 创建设备类型选择区域
        device_type_group = QGroupBox("设备类型")
        device_type_layout = QHBoxLayout()

        self.device_type_buttons = QButtonGroup(self)
        self.huawei_radio = QRadioButton("华为")
        self.h3c_radio = QRadioButton("H3C")
        self.huawei_radio.setChecked(True)

        self.huawei_radio.toggled.connect(self.parse_nat_config)
        self.h3c_radio.toggled.connect(self.parse_nat_config)

        self.device_type_buttons.addButton(self.huawei_radio)
        self.device_type_buttons.addButton(self.h3c_radio)

        device_type_layout.addWidget(self.huawei_radio)
        device_type_layout.addWidget(self.h3c_radio)
        device_type_layout.addStretch()

        device_type_group.setLayout(device_type_layout)
        layout.addWidget(device_type_group)

        # 创建输入区域
        input_group = QGroupBox("NAT Server配置")
        input_layout = QVBoxLayout()
        self.nat_input = QTextEdit()
        self.nat_input.setPlaceholderText("在此粘贴NAT Server配置命令...")
        self.nat_input.textChanged.connect(self.parse_nat_config)
        input_layout.addWidget(self.nat_input)
        input_group.setLayout(input_layout)
        layout.addWidget(input_group)

        # 创建结果显示区域
        result_group = QGroupBox("解析结果")
        result_layout = QVBoxLayout()

        # 添加导出按钮
        export_button = QPushButton("导出到Excel")
        export_button.clicked.connect(self.export_to_excel)
        result_layout.addWidget(export_button)

        # 创建结果表格
        self.result_table = QTableWidget()
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.result_table.setAlternatingRowColors(True)
        result_layout.addWidget(self.result_table)

        result_group.setLayout(result_layout)
        layout.addWidget(result_group)

    def parse_nat_config(self):
        """解析NAT配置并显示结果"""
        text = self.nat_input.toPlainText()
        if not text.strip():
            self.result_table.setRowCount(0)
            self.current_data = []
            self.failed_entries = []
            self.current_device_type = None
            return

        self.current_device_type = "huawei" if self.huawei_radio.isChecked() else "h3c"
        data, failed_entries = NATParser.parse_config(text, self.current_device_type)

        self.current_data = data
        self.failed_entries = failed_entries
        self.display_results(data, failed_entries)

    def display_results(self, data, failed_entries):
        """显示解析结果"""
        # 根据设备类型设置表头
        if self.current_device_type == "huawei":
            headers = ["名称", "协议", "全局IP", "全局端口", "内部IP", "内部端口"]
            keys = ['name', 'protocol', 'global_ip', 'global_port', 'inside_ip', 'inside_port']
        else:  # h3c
            headers = ["名称", "协议", "全局IP", "全局端口", "内部IP", "内部端口", "VRRP", "描述"]
            keys = ['rule', 'protocol', 'global_ip', 'global_port', 'inside_ip', 'inside_port', 'vrrp', 'description']

        self.result_table.setColumnCount(len(headers))
        self.result_table.setHorizontalHeaderLabels(headers)

        # 显示数据
        if data:
            self.result_table.setRowCount(len(data))
            for row, item in enumerate(data):
                # 显示所有字段
                for col, key in enumerate(keys):
                    self.result_table.setItem(row, col, QTableWidgetItem(str(item.get(key, '-'))))
        else:
            self.result_table.setRowCount(0)

        # 如果有解析失败的条目，添加到表格底部
        if failed_entries:
            start_row = self.result_table.rowCount()
            self.result_table.setRowCount(start_row + len(failed_entries))
            for i, entry in enumerate(failed_entries, start_row):
                self.result_table.setItem(i, 0, QTableWidgetItem("解析失败"))
                self.result_table.setItem(i, 1, QTableWidgetItem(entry))

    def export_to_excel(self):
        """导出数据到Excel文件"""
        if not self.current_data and not self.failed_entries:
            QMessageBox.warning(self, "导出警告", "没有可导出的数据！")
            return

        # 获取保存路径
        default_path = os.path.join(os.path.expanduser("~"), "Desktop", "映射表.xlsx")
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存Excel文件",
            default_path,
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "NAT Server配置"

            # 根据设备类型选择表头
            if self.current_device_type == "huawei":
                headers = ["名称", "协议", "全局IP", "全局端口", "内部IP", "内部端口", "原始命令"]
                keys = ['name', 'protocol', 'global_ip', 'global_port', 'inside_ip', 'inside_port', 'command']
            else:  # h3c
                headers = ["名称", "协议", "全局IP", "全局端口", "内部IP", "内部端口", "VRRP", "描述", "原始命令"]
                keys = ['rule', 'protocol', 'global_ip', 'global_port', 'inside_ip', 'inside_port', 'vrrp', 'description', 'command']

            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            # 写入成功解析的数据
            current_row = 2
            for item in self.current_data:
                for col, key in enumerate(keys, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = item.get(key, '-')
                    cell.alignment = Alignment(horizontal='left')
                    if key == 'command':
                        cell.alignment = Alignment(horizontal='left', wrap_text=True)
                current_row += 1

            # 写入解析失败的条目
            if self.failed_entries:
                # 添加一个空行
                current_row += 1

                # 添加"解析失败条目"标题行
                cell = ws.cell(row=current_row, column=1)
                cell.value = "解析失败条目："
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                current_row += 1

                # 写入失败条目
                for entry in self.failed_entries:
                    cell = ws.cell(row=current_row, column=1)
                    cell.value = entry
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
                    cell.alignment = Alignment(horizontal='left')
                    current_row += 1

            # 调整列宽
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # 为原始命令列单独设置更大宽度
            ws.column_dimensions[chr(ord('A') + len(headers) - 1)].width = 120

            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"数据已成功导出到:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "导出错误", f"导出过程中发生错误：\n{str(e)}")
