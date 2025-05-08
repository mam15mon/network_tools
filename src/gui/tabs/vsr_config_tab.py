from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, 
                           QPushButton, QTabWidget, QTextEdit,
                           QMessageBox, QFileDialog)
from PyQt6.QtCore import Qt
from jinja2 import Environment, FileSystemLoader
import os
import openpyxl
from datetime import datetime

class VSRConfigTab(QWidget):
    def __init__(self):
        super().__init__()
        self.setup_ui()
        
    def setup_ui(self):
        """设置UI界面"""
        layout = QVBoxLayout(self)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        
        # 模板导出按钮组
        template_group = QHBoxLayout()
        self.local_btn = QPushButton("本地模板")
        self.radius_btn = QPushButton("RADIUS模板") 
        self.ldap_btn = QPushButton("LDAP模板")
        
        self.local_btn.clicked.connect(lambda: self.export_template("local"))
        self.radius_btn.clicked.connect(lambda: self.export_template("radius"))
        self.ldap_btn.clicked.connect(lambda: self.export_template("ldap"))
        
        template_group.addWidget(self.local_btn)
        template_group.addWidget(self.radius_btn)
        template_group.addWidget(self.ldap_btn)
        
        # 导入按钮 - 绿色
        self.import_btn = QPushButton("导入Excel")
        self.import_btn.setProperty("class", "import-btn")
        
        # 生成配置按钮 - 橙色
        self.generate_btn = QPushButton("生成配置")
        self.generate_btn.setProperty("class", "generate-btn")
        self.import_btn.clicked.connect(self.import_excel)
        self.generate_btn.clicked.connect(self.generate_config)
        
        button_layout.addLayout(template_group)
        button_layout.addWidget(self.import_btn)
        button_layout.addWidget(self.generate_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # 使用QTabWidget来显示不同设备的配置
        self.config_tabs = QTabWidget()
        self.config_tabs.setTabPosition(QTabWidget.TabPosition.North)
        layout.addWidget(self.config_tabs)
        
    def export_template(self, template_type):
        """导出指定类型的Excel模板
        Args:
            template_type: 模板类型 - 'local'/'radius'/'ldap'
        """
        try:
            # 创建新的Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "VSR配置参数"
            
            # 基础表头
            headers = [
                "设备名称", "IP", "Gateway", 
                "VSR Username", "VSR Password",
                "Monitor Username", "Monitor Password", 
                "PPP Username", "PPP Password",
                "Start IP", "End IP", "Pool IP Gateway"
            ]
            
            # 根据模板类型添加特定表头
            if template_type == "radius":
                headers.extend([
                    "Radius IP", "Radius Password"
                ])
            elif template_type == "ldap":
                headers.extend([
                    "LDAP Server IP", "LDAP Login DN",
                    "LDAP Search Base DN", "LDAP Password",
                    "LDAP User Object Class", "LDAP Username Attribute"
                ])
            
            ws.append(headers)
            
            # 基础数据
            base_data = [
                "VSR-1", "192.168.1.1", "192.168.1.254",
                "admin", "admin123",
                "monitor", "monitor123",
                "ppp", "ppp123",
                "10.0.0.1", "10.0.0.254", "10.0.0.1"
            ]
            
            # 根据模板类型设置示例数据
            if template_type == "local":
                example_data = base_data + ["", "", "", "", "", ""]
            elif template_type == "radius":
                example_data = base_data + [
                    "192.168.100.200", "radius123",  # RADIUS参数
                    "", "", "", ""  # LDAP参数留空
                ]
            else:  # ldap
                example_data = base_data + [
                    "192.168.100.100",  # LDAP Server IP
                    "cn=admin,dc=example,dc=com",  # LDAP Login DN
                    "ou=users,dc=example,dc=com",  # LDAP Search Base DN
                    "ldap123",  # LDAP Password
                    "posixAccount",  # LDAP User Object Class
                    "uid"  # LDAP Username Attribute
                ]
            ws.append(example_data)
            
            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                ws.column_dimensions[column].width = max_length + 2
            
            # 保存文件
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "保存模板",
                "vsr_config_template.xlsx",
                "Excel文件 (*.xlsx)"
            )
            
            if file_name:
                wb.save(file_name)
                result = QMessageBox.information(
                    self,
                    "成功", 
                    f"{'本地' if template_type == 'local' else template_type.upper()}认证模板已导出",
                    QMessageBox.StandardButton.Ok
                )
                if result == QMessageBox.StandardButton.Ok:
                    try:
                        import os
                        os.startfile(file_name)
                    except Exception as e:
                        QMessageBox.warning(
                            self,
                            "警告",
                            f"无法自动打开文件: {str(e)}"
                        )
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出模板时出错：{str(e)}")
            
    def import_excel(self):
        """导入Excel文件"""
        try:
            file_name, _ = QFileDialog.getOpenFileName(
                self,
                "选择Excel文件",
                "",
                "Excel文件 (*.xlsx)"
            )
            
            if file_name:
                self.current_excel = file_name
                self.preview_excel(file_name)
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入Excel时出错：{str(e)}")
            
    def preview_excel(self, file_name):
        """预览Excel内容"""
        try:
            wb = openpyxl.load_workbook(file_name)
            ws = wb.active
            
            # 清除现有的标签页
            while self.config_tabs.count():
                self.config_tabs.removeTab(0)
            
            # 读取数据
            self.excel_data = []
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # 跳过空行
                    device_data = dict(zip(headers, row))
                    self.excel_data.append(device_data)
                    
                    # 为每个设备创建一个标签页
                    device_name = device_data.get("设备名称", f"设备-{len(self.excel_data)}")
                    device_tab = QTextEdit()
                    device_tab.setReadOnly(True)
                    device_tab.setFont(self.get_monospace_font())
                    device_tab.setText("点击生成配置查看")
                    
                    self.config_tabs.addTab(device_tab, device_name)
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"预览Excel时出错：{str(e)}")
    
    def get_monospace_font(self):
        """获取等宽字体"""
        font = self.font()
        font.setFamily("Courier New")  # 使用等宽字体
        font.setPointSize(10)
        return font
            
    def generate_config(self):
        """生成VSR配置"""
        try:
            if not hasattr(self, 'excel_data'):
                QMessageBox.warning(self, "警告", "请先导入Excel文件")
                return
                
            # 获取模板
            template_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'templates')
            env = Environment(loader=FileSystemLoader(template_dir))
            template = env.get_template('vsr_config.j2')
            
            # 创建导出目录
            export_dir = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'export', 'vsr配置')
            os.makedirs(export_dir, exist_ok=True)
            
            # 生成时间戳
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # 为每个设备生成配置
            for i, data in enumerate(self.excel_data):
                # 准备模板数据
                template_data = {
                    'ip': data.get('IP'),
                    'gateway': data.get('Gateway'),
                    'vsr_username': data.get('VSR Username'),
                    'vsr_password': data.get('VSR Password'),
                    'monitor_username': data.get('Monitor Username'),
                    'monitor_password': data.get('Monitor Password'),
                    'ppp_username': data.get('PPP Username'),
                    'ppp_password': data.get('PPP Password'),
                    'start_ip': data.get('Start IP'),
                    'end_ip': data.get('End IP'),
                    'pool_ip_gateway': data.get('Pool IP Gateway'),
                    'radius_ip': data.get('Radius IP'),
                    'radius_password': data.get('Radius Password'),
                    # LDAP参数
                    'ldap_server_ip': data.get('LDAP Server IP'),
                    'ldap_login_dn': data.get('LDAP Login DN'),
                    'ldap_search_base_dn': data.get('LDAP Search Base DN'),
                    'ldap_password': data.get('LDAP Password'),
                    'ldap_user_object_class': data.get('LDAP User Object Class'),
                    'ldap_username_attribute': data.get('LDAP Username Attribute')
                }
                
                # 渲染配置
                config = template.render(template_data)
                
                # 更新标签页内容
                device_tab = self.config_tabs.widget(i)
                if device_tab:
                    device_tab.setText(config)
                
                # 保存配置文件
                device_name = data.get("设备名称", f"设备-{i+1}")
                file_name = f"{device_name}_{timestamp}.txt"
                file_path = os.path.join(export_dir, file_name)
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(config)
            
            QMessageBox.information(self, "成功", f"配置已生成并保存到：{export_dir}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成配置时出错：{str(e)}")
